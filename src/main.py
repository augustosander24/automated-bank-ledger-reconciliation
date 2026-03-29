from datetime import date, datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import (
    ensure_directories,
    BANK_CLEAN_FILE,
    LEDGER_CLEAN_FILE,
    OBLIGATIONS_CLEAN_FILE,
)
from ingestion import load_raw_sources
from cleaning import (
    clean_bank_transactions,
    clean_ledger_transactions,
    clean_payment_obligations,
)
from reconciliation import (
    build_final_output_views,
    reconcile_bank_vs_ledger,
    save_final_output_views,
)


HEADER_OVERRIDES = {
    "metric": "Metric",
    "value": "Value",
    "note": "Note",
    "match_id": "Match ID",
    "match_type": "Match Type",
    "match_status": "Match Status",
    "bank_txn_id": "Bank Transaction ID",
    "ledger_txn_id": "Ledger Transaction ID",
    "bank_date": "Bank Date",
    "ledger_date": "Ledger Date",
    "transaction_date": "Transaction Date",
    "posting_date": "Posting Date",
    "bank_amount": "Bank Amount",
    "ledger_amount": "Ledger Amount",
    "bank_cash_effect": "Bank Cash Effect",
    "ledger_cash_effect": "Ledger Cash Effect",
    "amount_difference": "Amount Difference",
    "amount_variance": "Amount Variance",
    "date_difference_days": "Date Difference (Days)",
    "bank_reference": "Bank Reference",
    "ledger_reference": "Ledger Reference",
    "bank_description": "Bank Description",
    "ledger_description": "Ledger Description",
    "bank_debit_credit": "Bank Debit / Credit",
    "ledger_debit_credit": "Ledger Debit / Credit",
    "counterparty_name": "Counterparty Name",
    "vendor_customer_name": "Vendor / Customer Name",
    "exception_id": "Exception ID",
    "exception_code": "Exception Code",
    "exception_reason": "Exception Reason",
    "review_status": "Review Status",
    "journal_source": "Journal Source",
    "gl_account": "GL Account",
    "debit_credit": "Debit / Credit",
    "bank_working_status": "Bank Working Status",
    "ledger_working_status": "Ledger Working Status",
}

DATE_COLUMN_KEYWORDS = (
    "date",
)

DECIMAL_COLUMN_KEYWORDS = (
    "amount",
    "balance",
    "variance",
    "difference",
    "delta",
    "value",
)

INTEGER_COLUMN_KEYWORDS = (
    "count",
    "days",
    "qty",
    "quantity",
    "rows",
)

TABLE_STYLE_NAME = "TableStyleMedium2"
HEADER_FILL = "1F4E78"
HEADER_TEXT = "FFFFFF"
BORDER_COLOR = "D9D9D9"

DECIMAL_NUMBER_FORMAT = '#,##0.00;-#,##0.00'
INTEGER_NUMBER_FORMAT = '#,##0'
DATE_NUMBER_FORMAT = 'yyyy-mm-dd'
TEXT_FORMAT = '@'

TEXT_COLUMN_NAMES = {
    "match id",
    "bank transaction id",
    "ledger transaction id",
    "exception id",
    "source record id",
    "bank reference number",
    "ledger journal reference",
    "ledger source reference",
    "reference number",
    "source reference",
    "gl account number",
    "gl account name",
}

WRAP_TEXT_COLUMN_KEYWORDS = (
    "description",
    "note",
    "action",
    "reason",
)

WIDE_TEXT_COLUMN_NAMES = {
    "exception category description",
    "bank description",
    "ledger description",
    "recommended action",
    "notes",
    "reconciliation exclusion reason",
}

BALANCE_COLUMN_NAMES = {
    "balance after transaction",
    "balance after posting",
    "adjusted bank balance",
    "adjusted ledger balance",
}


def prettify_header(column_name: str) -> str:
    lower_name = column_name.lower()
    if lower_name in HEADER_OVERRIDES:
        return HEADER_OVERRIDES[lower_name]

    words = column_name.replace("_", " ").split()
    pretty_words = []

    for word in words:
        upper_word = word.upper()
        if upper_word in {"ID", "GL", "AP", "AR", "ERP"}:
            pretty_words.append(upper_word)
        else:
            pretty_words.append(word.capitalize())

    return " ".join(pretty_words)


def convert_numeric_like_text(value):
    if pd.isna(value):
        return value

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value

    text_value = str(value).strip()
    if text_value == "":
        return value

    cleaned_value = text_value.replace(",", "").replace("$", "")

    if cleaned_value.startswith("(") and cleaned_value.endswith(")"):
        cleaned_value = f"-{cleaned_value[1:-1]}"

    try:
        numeric_value = float(cleaned_value)
    except ValueError:
        return value

    if numeric_value.is_integer():
        return int(numeric_value)

    return numeric_value


def convert_date_like_text(value):
    if pd.isna(value):
        return value

    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value

    text_value = str(value).strip()
    if text_value == "":
        return value

    parsed_value = pd.to_datetime(text_value, errors="coerce")
    if pd.isna(parsed_value):
        return value

    return parsed_value.to_pydatetime()


def is_decimal_column(column_name: str) -> bool:
    lower_name = column_name.lower()
    return any(keyword in lower_name for keyword in DECIMAL_COLUMN_KEYWORDS)


def is_integer_column(column_name: str) -> bool:
    lower_name = column_name.lower()
    return any(keyword in lower_name for keyword in INTEGER_COLUMN_KEYWORDS)


def is_date_column(column_name: str) -> bool:
    lower_name = column_name.lower()
    return any(keyword in lower_name for keyword in DATE_COLUMN_KEYWORDS)


def is_text_column(column_name: str) -> bool:
    return column_name.lower() in TEXT_COLUMN_NAMES


def should_wrap_column(column_name: str) -> bool:
    lower_name = column_name.lower()
    return any(keyword in lower_name for keyword in WRAP_TEXT_COLUMN_KEYWORDS)


def get_column_width(column_name: str, max_length: int) -> int:
    lower_name = column_name.lower()

    if lower_name in WIDE_TEXT_COLUMN_NAMES:
        return min(max(max_length + 2, 28), 60)
    if should_wrap_column(column_name):
        return min(max(max_length + 2, 24), 48)
    if is_date_column(column_name):
        return max(14, min(max_length + 2, 18))
    if is_decimal_column(column_name) or lower_name in BALANCE_COLUMN_NAMES:
        return max(14, min(max_length + 2, 18))
    if is_text_column(column_name):
        return max(16, min(max_length + 2, 24))

    return min(max_length + 2, 36)


def prepare_dataframe_for_excel(dataframe: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    prepared_df = dataframe.copy()

    for column_name in prepared_df.columns:
        lower_name = column_name.lower()

        if is_date_column(column_name):
            prepared_df[column_name] = prepared_df[column_name].apply(convert_date_like_text)
            continue

        if is_text_column(column_name):
            prepared_df[column_name] = prepared_df[column_name].where(
                prepared_df[column_name].isna(),
                prepared_df[column_name].astype(str),
            )
            continue

        if sheet_name == "Reconciliation Summary" and lower_name == "value":
            prepared_df[column_name] = prepared_df[column_name].apply(convert_numeric_like_text)
            continue

        if is_decimal_column(column_name) or is_integer_column(column_name):
            prepared_df[column_name] = prepared_df[column_name].apply(convert_numeric_like_text)

    return prepared_df


def apply_header_style(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    header_font = Font(color=HEADER_TEXT, bold=True)
    header_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    worksheet.row_dimensions[1].height = 22

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border


def auto_fit_and_format_columns(
    worksheet,
    display_df: pd.DataFrame,
    original_columns: list[str],
    sheet_name: str,
) -> None:
    for col_idx, original_name in enumerate(original_columns, start=1):
        column_letter = get_column_letter(col_idx)
        display_name = str(display_df.columns[col_idx - 1])
        max_length = len(display_name)

        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet[f"{column_letter}{row_idx}"]
            cell_value = cell.value
            wrap_text = should_wrap_column(original_name)

            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))

            if is_date_column(original_name):
                if isinstance(cell_value, (datetime, date)):
                    cell.number_format = DATE_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)
            elif is_text_column(original_name):
                cell.number_format = TEXT_FORMAT
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap_text)
            elif is_decimal_column(original_name):
                cell.number_format = DECIMAL_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap_text)
            elif is_integer_column(original_name):
                cell.number_format = INTEGER_NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=wrap_text)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap_text)

        if sheet_name == "Reconciliation Summary" and original_name.lower() == "metric":
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet[f"{column_letter}{row_idx}"].font = Font(bold=True)

        worksheet.column_dimensions[column_letter].width = get_column_width(original_name, max_length)


def add_excel_table(worksheet, sheet_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    table_ref = worksheet.dimensions
    safe_name = re.sub(r"[^A-Za-z0-9]", "", sheet_name)
    table_name = f"Tbl{safe_name}"

    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def format_worksheet(
    worksheet,
    display_df: pd.DataFrame,
    original_columns: list[str],
    sheet_name: str,
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    apply_header_style(worksheet)
    auto_fit_and_format_columns(
        worksheet=worksheet,
        display_df=display_df,
        original_columns=original_columns,
        sheet_name=sheet_name,
    )
    add_excel_table(worksheet=worksheet, sheet_name=sheet_name)


def build_output_workbook(
    output_views: dict[str, pd.DataFrame],
    output_dir: str = "output",
    workbook_name: str = "reconciliation_output.xlsx",
) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    workbook_path = output_path / workbook_name
    fallback_workbook_path = output_path / f"{workbook_path.stem}_rebuilt{workbook_path.suffix}"
    sheet_order = [
        "Reconciliation Summary",
        "Reconciled Matches",
        "Exceptions Report",
        "Legend",
        "Clean Bank Data",
        "Clean Ledger Data",
    ]

    def write_workbook(target_path: Path) -> Path:
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            for sheet_name in sheet_order:
                raw_df = output_views[sheet_name].copy()
                prepared_df = prepare_dataframe_for_excel(raw_df, sheet_name)

                display_df = prepared_df.copy()
                original_columns = list(display_df.columns)
                display_df.columns = [prettify_header(column_name) for column_name in original_columns]

                display_df.to_excel(writer, sheet_name=sheet_name, index=False)

                worksheet = writer.sheets[sheet_name]
                format_worksheet(
                    worksheet=worksheet,
                    display_df=display_df,
                    original_columns=original_columns,
                    sheet_name=sheet_name,
                )

        return target_path

    try:
        return write_workbook(workbook_path)
    except PermissionError:
        return write_workbook(fallback_workbook_path)


def main() -> None:
    ensure_directories()

    sources = load_raw_sources()

    bank_cleaned = clean_bank_transactions(sources["raw_bank"])
    ledger_cleaned = clean_ledger_transactions(sources["raw_ledger"])
    obligations_cleaned = clean_payment_obligations(sources["raw_obligations"])

    bank_cleaned.to_csv(BANK_CLEAN_FILE, index=False)
    ledger_cleaned.to_csv(LEDGER_CLEAN_FILE, index=False)
    obligations_cleaned.to_csv(OBLIGATIONS_CLEAN_FILE, index=False)

    reconciliation_result = reconcile_bank_vs_ledger(
        bank_df=bank_cleaned,
        ledger_df=ledger_cleaned,
    )

    output_views = build_final_output_views(
        result=reconciliation_result,
    )

    save_final_output_views(
        output_views=output_views,
        output_dir="output",
    )

    workbook_path = build_output_workbook(
        output_views=output_views,
        output_dir="output",
        workbook_name="reconciliation_output.xlsx",
    )

    print("Milestone 1 complete.")
    print(f"Bank cleaned output: {BANK_CLEAN_FILE}")
    print(f"Ledger cleaned output: {LEDGER_CLEAN_FILE}")
    print(f"Obligations cleaned output: {OBLIGATIONS_CLEAN_FILE}")

    print("\nRow counts:")
    print(f"Bank cleaned rows: {len(bank_cleaned)}")
    print(f"Ledger cleaned rows: {len(ledger_cleaned)}")
    print(f"Obligations cleaned rows: {len(obligations_cleaned)}")

    print("\nReconciliation complete.")
    print("Reconciliation outputs saved to: output")
    print(f"Excel workbook saved to: {workbook_path}")


if __name__ == "__main__":
    main()
